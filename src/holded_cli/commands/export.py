from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import typer
from rich.text import Text

from holded_cli.console import get_output_console
from holded_cli.dates import parse_date
from holded_cli.errors import InputError
from holded_cli.holded_client import HoldedClient
from holded_cli.state import AppState


_ES_MONTHS = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

_APPROVED_STATUS = {
    None: "Pendiente",
    "approved": "Aprobado",
    "rejected": "Rechazado",
}


def _utc_to_local_hhmm(utc_iso: str, tz_name: str) -> str:
    dt = datetime.fromisoformat(utc_iso)
    return dt.astimezone(ZoneInfo(tz_name)).strftime("%H:%M")


def _fmt_duration(seconds: int) -> str:
    """Format seconds as 'NNh MMm', or empty string if zero."""
    if not seconds:
        return ""
    h, m = divmod(abs(seconds) // 60, 60)
    return f"{h:02d}h {m:02d}m"


def _build_xlsx(
    data: list[dict],
    tz_name: str,
    from_date: datetime,
    to_date: datetime,
    workplace_map: dict[str, str],
    employee_name: str,
    company_name: str,
) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    NUM_COLS = 8
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro horario"

    def _merge_bold(row: int, value: str, size: int = 11) -> None:
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws.cell(row=row, column=1, value=value)
        cell.font = Font(bold=True, size=size)

    # --- Header block ---
    if company_name:
        _merge_bold(1, company_name, size=12)
    _merge_bold(2, employee_name)

    ws.append([])  # row 3 empty

    # Title: "Registros de control horario - Mes YYYY"
    month_str = _ES_MONTHS.get(from_date.month, "") if from_date.month == to_date.month else ""
    if month_str:
        title = f"Registros de control horario - {month_str} {from_date.year}"
    else:
        title = f"Registros de control horario - {from_date.strftime('%d/%m/%Y')} - {to_date.strftime('%d/%m/%Y')}"
    _merge_bold(4, title, size=12)

    ws.append([])  # row 5 empty

    # --- Column headers ---
    headers = [
        "Día", "Horario", "Horas totales", "Total horas trabajadas",
        "Total horas pausadas", "Horas planeadas", "Ubicación", "Estado",
    ]
    ws.append(headers)
    header_row = ws.max_row
    header_fill = PatternFill("solid", fgColor="1F3864")
    center = Alignment(horizontal="center", vertical="center")
    for col in range(1, NUM_COLS + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[header_row].height = 18

    # --- Data rows ---
    dim_font = Font(color="AAAAAA", size=10)
    normal_font = Font(size=10)

    for entry in data:
        d = entry["date"]
        dt = datetime.fromisoformat(d)
        is_weekend = dt.weekday() >= 5
        trackers = entry.get("trackers", [])
        timeoffs = entry.get("timeoffs", [])
        stats = entry.get("stats", {})

        day_str = dt.strftime("%d/%m/%Y")

        if is_weekend:
            ws.append([day_str, "", "", "", "", "", "", ""])
            row_idx = ws.max_row
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=row_idx, column=col).font = dim_font
        elif timeoffs:
            holiday_name = timeoffs[0].get("name", "Festivo")
            ws.append([day_str, holiday_name, "", "", "", "", "", ""])
            row_idx = ws.max_row
            for col in range(1, NUM_COLS + 1):
                c = ws.cell(row=row_idx, column=col)
                c.font = Font(italic=True, color="666666", size=10)
        elif trackers:
            tracker = trackers[0]
            start = _utc_to_local_hhmm(tracker["start"], tz_name)
            end = _utc_to_local_hhmm(tracker["end"], tz_name)
            horario = f"{start} - {end}"
            total = _fmt_duration(tracker.get("time", 0) or tracker.get("duration", 0))
            worked = _fmt_duration(tracker.get("effectiveWorkedTime", 0))
            paused = _fmt_duration(tracker.get("pausedTime", 0))
            expected = _fmt_duration(stats.get("expectedTime", 0))
            workplace_id = tracker.get("workplaceId") or ""
            location = workplace_map.get(workplace_id, "")
            status = _APPROVED_STATUS.get(tracker.get("approvedStatus"), "Pendiente")
            ws.append([day_str, horario, total, worked, paused, expected, location, status])
            row_idx = ws.max_row
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=row_idx, column=col).font = normal_font
        else:
            expected = _fmt_duration(stats.get("expectedTime", 0))
            ws.append([day_str, "", "", "", "", expected, "", ""])
            row_idx = ws.max_row
            for col in range(1, NUM_COLS + 1):
                ws.cell(row=row_idx, column=col).font = dim_font

        # subtle bottom border on every data row
        row_idx = ws.max_row
        for col in range(1, NUM_COLS + 1):
            ws.cell(row=row_idx, column=col).border = border

    # --- Footer ---
    ws.append([])
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.append([f"Informe creado automáticamente con Holded - {now_str}"])
    footer_row = ws.max_row
    ws.cell(row=footer_row, column=1).font = Font(italic=True, color="999999", size=9)

    # --- Column widths ---
    col_widths = [13, 18, 15, 22, 22, 15, 16, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_command(
    ctx: typer.Context,
    from_date: str = typer.Option(..., "--from", help="Start date in YYYY-MM-DD format."),
    to_date: str = typer.Option(..., "--to", help="End date in YYYY-MM-DD format."),
    fmt: str = typer.Option("pdf", "--format", help="Output format: pdf or xlsx."),
    out: Optional[Path] = typer.Option(
        None, "--out", help="Output file path. Defaults to current directory."
    ),
    company: Optional[str] = typer.Option(
        None, "--company", help="Company name for the Excel header."
    ),
) -> None:
    """Export time-tracking records as PDF or Excel for a date range.

    Example:
      holded export --from 2026-04-01 --to 2026-04-30
      holded export --from 2026-04-01 --to 2026-04-30 --format xlsx
      holded export --from 2026-04-01 --to 2026-04-30 --format xlsx --company 'ACME S.L.'
    """
    state: AppState = ctx.obj

    if fmt not in ("pdf", "xlsx"):
        raise InputError(
            message=f"Unknown format: {fmt!r}",
            hint="Use --format pdf or --format xlsx.",
        )

    resolved_from = parse_date(from_date)
    resolved_to = parse_date(to_date)

    if resolved_from > resolved_to:
        raise InputError(
            message=f"Start date {resolved_from} is after end date {resolved_to}.",
            hint="The --from date must be on or before --to.",
        )

    if out is None:
        filename = f"holded-{resolved_from}_{resolved_to}.{fmt}"
        out = Path.cwd() / filename

    console = get_output_console()

    with HoldedClient(state.session_store) as client:
        if fmt == "pdf":
            console.print("[dim]Fetching PDF…[/dim]")
            content = client.get_timetracking_pdf(
                resolved_from, resolved_to, state.config.timezone
            )
        else:
            console.print("[dim]Fetching data…[/dim]")
            data = client.get_timetracking_data(
                resolved_from, resolved_to, state.config.timezone
            )
            workplaces = client.get_workplaces()
            workplace_map = {wp["id"]: wp["name"] for wp in workplaces}

            employee_name = ""
            for entry in data:
                for tracker in entry.get("trackers", []):
                    employee_name = tracker.get("employeeName", "")
                    if employee_name:
                        break
                if employee_name:
                    break

            content = _build_xlsx(
                data=data,
                tz_name=state.config.timezone,
                from_date=datetime.fromisoformat(str(resolved_from)),
                to_date=datetime.fromisoformat(str(resolved_to)),
                workplace_map=workplace_map,
                employee_name=employee_name,
                company_name=company or "",
            )

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(content)

    result = Text()
    result.append("✓  ", style="green bold")
    result.append(str(out), style="bold")
    result.append(f"  ·  {len(content) / 1024:.1f} KB", style="dim")
    console.print(result)
